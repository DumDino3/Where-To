"""
Export each sheet in an Excel workbook to a CSV file.

Usage:
  python export_excel_sheets_to_csv.py path/to/data.xlsx --out ./Exports

Options:
  --all                 Export all sheets (default)
  --sheets ride_request npc locations dialogue_set
  --skip-hidden         Skip hidden sheets (default: true)
  --keep-empty-rows     Keep empty rows (default: false)
  --keep-empty-cols     Keep empty columns (default: true)
  --encoding utf-8-sig  (default: utf-8-sig, good for Excel/Unity)
"""

import argparse
import csv
import glob
import os
from typing import List, Optional

from openpyxl import load_workbook


def safe_filename(name: str) -> str:
  # Make a filesystem-safe filename based on sheet name
  keep = []
  for ch in name.strip():
    if ch.isalnum() or ch in ("-", "_", "."):
      keep.append(ch)
    elif ch.isspace():
      keep.append("_")
    else:
      keep.append("_")
  out = "".join(keep)
  return out if out else "sheet"


def is_row_empty(row) -> bool:
  return all(cell is None or (isinstance(cell, str) and cell.strip() == "") for cell in row)


def clear_csv_files(out_dir: str):
  """Delete all .csv files in the output directory."""
  if os.path.exists(out_dir):
    csv_files = glob.glob(os.path.join(out_dir, "*.csv"))
    for csv_file in csv_files:
      try:
        os.remove(csv_file)
        print(f"Deleted: {csv_file}")
      except Exception as e:
        print(f"Warning: Could not delete {csv_file}: {e}")


def export_sheet_to_csv(ws, out_path: str, *, keep_empty_rows: bool, keep_empty_cols: bool, encoding: str):
  # Read all rows up to ws.max_row/ws.max_column
  max_row = ws.max_row
  max_col = ws.max_column

  rows = []
  # Start from row 2 to skip the first row
  for r in range(2, max_row + 1):
    row = []
    # Start from column 1 to include the first column
    for c in range(1, max_col + 1):
      v = ws.cell(row=r, column=c).value
      # Normalize newlines, keep numbers/bools as-is (csv module will stringify)
      if isinstance(v, str):
        v = v.replace("\r\n", "\n").replace("\r", "\n").strip()
      row.append(v)
    rows.append(row)

  if not keep_empty_rows:
    rows = [r for r in rows if not is_row_empty(r)]

  if not keep_empty_cols and rows:
    # Drop columns that are entirely empty
    col_count = max(len(r) for r in rows) if rows else 0
    if col_count > 0:
      keep = [False] * col_count
      for r in rows:
        for i, v in enumerate(r):
          if v is not None and not (isinstance(v, str) and v.strip() == ""):
            keep[i] = True
      rows = [[v for i, v in enumerate(r) if keep[i]] for r in rows]

  os.makedirs(os.path.dirname(out_path), exist_ok=True)
  with open(out_path, "w", newline="", encoding=encoding) as f:
    writer = csv.writer(f, delimiter=",", quotechar='"', quoting=csv.QUOTE_MINIMAL)
    for r in rows:
      writer.writerow(["" if v is None else v for v in r])


def export_workbook(
  xlsx_path: str,
  out_dir: str,
  *,
  sheet_names: Optional[List[str]] = None,
  export_all: bool = True,
  skip_hidden: bool = True,
  keep_empty_rows: bool = False,
  keep_empty_cols: bool = True,
  encoding: str = "utf-8-sig"
):
  # Clear old CSV files before exporting
  clear_csv_files(out_dir)
  
  wb = load_workbook(filename=xlsx_path, data_only=True)

  if sheet_names and not export_all:
    targets = sheet_names
  else:
    targets = wb.sheetnames

  exported = []
  for name in targets:
    if name not in wb.sheetnames:
      raise ValueError(f'Sheet "{name}" not found in workbook. Available: {wb.sheetnames}')

    ws = wb[name]
    if skip_hidden and ws.sheet_state != "visible":
      continue

    csv_name = safe_filename(name) + ".csv"
    out_path = os.path.join(out_dir, csv_name)
    export_sheet_to_csv(ws, out_path, keep_empty_rows=keep_empty_rows, keep_empty_cols=keep_empty_cols, encoding=encoding)
    exported.append(out_path)

  return exported


def main():
  script_dir = os.path.dirname(os.path.abspath(__file__))
  project_root = os.path.normpath(os.path.join(script_dir, "..", "..", ".."))
  default_out = os.path.normpath(os.path.join(project_root, "Assets", "Resources", "CsvDatabase"))

  def resolve_path(path: str) -> str:
    if os.path.isabs(path):
      return os.path.abspath(path)

    cwd_path = os.path.abspath(os.path.join(os.getcwd(), path))
    if os.path.exists(cwd_path):
      return cwd_path

    script_path = os.path.abspath(os.path.join(script_dir, path))
    if os.path.exists(script_path):
      return script_path

    # If the user provided a Resources-based relative path, resolve from the project Assets/Resources root.
    normalized = path.replace("/", os.sep).replace("\\", os.sep)
    if os.sep + "Resources" + os.sep in os.sep + normalized:
      suffix = normalized.split(os.sep + "Resources" + os.sep, 1)[1]
      return os.path.abspath(os.path.join(project_root, "Assets", "Resources", suffix))

    return script_path

  ap = argparse.ArgumentParser()
  ap.add_argument("xlsx", help="Path to .xlsx workbook")
  ap.add_argument("--out", default=default_out, help="Output directory for CSV files")

  group = ap.add_mutually_exclusive_group()
  group.add_argument("--all", action="store_true", help="Export all sheets (default)")
  group.add_argument("--sheets", nargs="+", help="Specific sheet names to export")

  ap.add_argument("--skip-hidden", action="store_true", default=True, help="Skip hidden sheets (default: true)")
  ap.add_argument("--include-hidden", dest="skip_hidden", action="store_false", help="Include hidden sheets")
  ap.add_argument("--keep-empty-rows", action="store_true", help="Keep empty rows (default: false)")
  ap.add_argument("--drop-empty-cols", action="store_true", help="Drop completely empty columns (default: keep)")
  ap.add_argument("--encoding", default="utf-8-sig", help="Output encoding (default: utf-8-sig)")

  args = ap.parse_args()

  xlsx_path = resolve_path(args.xlsx)
  out_dir = resolve_path(args.out)

  export_all = True
  sheet_names = None
  if args.sheets:
    export_all = False
    sheet_names = args.sheets

  exported = export_workbook(
    xlsx_path,
    out_dir,
    sheet_names=sheet_names,
    export_all=export_all,
    skip_hidden=args.skip_hidden,
    keep_empty_rows=args.keep_empty_rows,
    keep_empty_cols=not args.drop_empty_cols,
    encoding=args.encoding,
  )

  print("Exported:")
  for p in exported:
    print(" -", p)


if __name__ == "__main__":
  main()